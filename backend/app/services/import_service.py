"""
services/import_service.py — Servicio de Importación Masiva de Productos desde Excel
===================================================================================

¿QUÉ HACE ESTE ARCHIVO? (Explicado para dummies 💡)
---------------------------------------------------
Imagina que te entregan una hoja de cálculo Excel (.xlsx) con 100 productos.
Algunos son nuevos, otros ya los vendes (pero cambiaron de precio), y otros fueron
escritos con errores humanos (ej: precio = "gratis", o stock = -5).

Este servicio hace el trabajo de un auditor inteligente (Decisión [DEC-008]):
1. Abre el archivo Excel directamente desde la memoria (sin guardarlo en disco).
2. Lee los encabezados para saber qué columna es cada dato.
3. Revisa fila por fila:
   - Si una fila tiene error -> La anota en la lista de errores con su número de fila y campo,
     pero NO detiene el proceso (Resiliencia Parcial).
   - Si el SKU ya está dos veces dentro del mismo archivo Excel -> Lo rechaza por duplicidad.
   - Si el SKU ya existía en la base de datos -> ACTUALIZA los datos (UPDATE).
   - Si el SKU es nuevo -> CREA el producto en la base de datos (INSERT).
4. Devuelve un informe claro:
   {
     "leidas": 50,
     "insertadas": 45,
     "actualizadas": 3,
     "rechazadas": 2,
     "errores": [...]
   }
"""

import io
from decimal import Decimal, InvalidOperation
from typing import BinaryIO, Dict, List, Set
import openpyxl
from sqlmodel import Session, select

from app.models.product import Product
from app.schemas.import_result import ImportResult, RowError


class ImportService:
    """
    Servicio para procesamiento de archivos Excel de catálogo con resiliencia parcial.
    """

    REQUIRED_COLUMNS = {"sku", "nombre", "categoria", "precio", "stock"}

    @classmethod
    def process_excel(cls, session: Session, file_content: BinaryIO) -> ImportResult:
        """
        Procesa el contenido binario de un archivo .xlsx y ejecuta las inserciones
        o actualizaciones pertinentes en la base de datos.
        """
        try:
            workbook = openpyxl.load_workbook(filename=io.BytesIO(file_content.read()), data_only=True)
            sheet = workbook.active
        except Exception as exc:
            return ImportResult(
                leidas=0,
                insertadas=0,
                actualizadas=0,
                rechazadas=0,
                errores=[
                    RowError(
                        fila=0,
                        campo="archivo",
                        motivo=f"No se pudo leer el archivo Excel. Asegúrate de subir un archivo .xlsx válido. ({str(exc)})",
                    )
                ],
            )

        # 1. Mapear encabezados de la fila 1
        headers: Dict[str, int] = {}
        first_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        if not first_row or not first_row[0]:
            return ImportResult(
                leidas=0,
                insertadas=0,
                actualizadas=0,
                rechazadas=0,
                errores=[RowError(fila=1, campo="encabezados", motivo="El archivo Excel está vacío.")],
            )

        for col_idx, col_name in enumerate(first_row[0]):
            if col_name is not None:
                clean_name = str(col_name).strip().lower()
                headers[clean_name] = col_idx

        # Validar columnas requeridas
        missing_cols = cls.REQUIRED_COLUMNS - set(headers.keys())
        if missing_cols:
            return ImportResult(
                leidas=0,
                insertadas=0,
                actualizadas=0,
                rechazadas=0,
                errores=[
                    RowError(
                        fila=1,
                        campo="encabezados",
                        motivo=f"Faltan las siguientes columnas obligatorias: {', '.join(sorted(missing_cols))}",
                    )
                ],
            )

        # 2. Procesar filas de datos a partir de la fila 2
        total_leidas = 0
        insertadas = 0
        actualizadas = 0
        rechazadas = 0
        errores: List[RowError] = []

        # Rastrear SKUs vistos en este archivo para evitar orden-dependencia y duplicados internos
        skus_in_file: Set[str] = set()

        rows = list(sheet.iter_rows(min_row=2, values_only=True))

        for row_index, row in enumerate(rows, start=2):
            # Saltar filas completamente vacías
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            total_leidas += 1

            # Extraer valores de celda de forma segura
            def get_val(col_name: str):
                idx = headers.get(col_name)
                if idx is not None and idx < len(row):
                    v = row[idx]
                    return v if v is not None else None
                return None

            raw_sku = get_val("sku")
            raw_nombre = get_val("nombre")
            raw_categoria = get_val("categoria")
            raw_precio = get_val("precio")
            raw_stock = get_val("stock")
            raw_imagen_url = get_val("imagen_url")
            raw_descripcion = get_val("descripcion")

            # Validar SKU
            if raw_sku is None or not str(raw_sku).strip():
                errores.append(RowError(fila=row_index, campo="sku", motivo="El SKU es obligatorio y no puede estar vacío."))
                rechazadas += 1
                continue

            sku = str(raw_sku).strip()

            # Validar SKU duplicado dentro del archivo Excel (DEC-008)
            if sku in skus_in_file:
                errores.append(
                    RowError(
                        fila=row_index,
                        campo="sku",
                        motivo=f"SKU '{sku}' duplicado dentro del mismo archivo Excel.",
                    )
                )
                rechazadas += 1
                continue

            skus_in_file.add(sku)

            # Validar Nombre
            if raw_nombre is None or not str(raw_nombre).strip():
                errores.append(RowError(fila=row_index, campo="nombre", motivo="El nombre del producto es obligatorio."))
                rechazadas += 1
                continue
            nombre = str(raw_nombre).strip()

            # Validar Categoría
            if raw_categoria is None or not str(raw_categoria).strip():
                errores.append(RowError(fila=row_index, campo="categoria", motivo="La categoría es obligatoria."))
                rechazadas += 1
                continue
            categoria = str(raw_categoria).strip()

            # Validar Precio
            if raw_precio is None:
                errores.append(RowError(fila=row_index, campo="precio", motivo="El precio es obligatorio."))
                rechazadas += 1
                continue
            try:
                precio_dec = Decimal(str(raw_precio).strip().replace("$", "").replace(",", "."))
                if precio_dec <= Decimal("0"):
                    raise ValueError()
            except (InvalidOperation, ValueError):
                errores.append(
                    RowError(
                        fila=row_index,
                        campo="precio",
                        motivo=f"Precio inválido ('{raw_precio}'). Debe ser un valor numérico mayor a 0.",
                    )
                )
                rechazadas += 1
                continue

            # Validar Stock
            if raw_stock is None:
                stock_int = 0
            else:
                try:
                    stock_int = int(float(str(raw_stock).strip()))
                    if stock_int < 0:
                        raise ValueError()
                except (ValueError, TypeError):
                    errores.append(
                        RowError(
                            fila=row_index,
                            campo="stock",
                            motivo=f"Stock inválido ('{raw_stock}'). Debe ser un número entero mayor o igual a 0.",
                        )
                    )
                    rechazadas += 1
                    continue

            imagen_url = str(raw_imagen_url).strip() if raw_imagen_url else None
            descripcion = str(raw_descripcion).strip() if raw_descripcion else None

            # 3. Consultar si el SKU ya existe en la Base de Datos
            existing_product = session.exec(select(Product).where(Product.sku == sku)).first()

            if existing_product:
                # ── UPDATE
                existing_product.nombre = nombre
                existing_product.categoria = categoria
                existing_product.precio = precio_dec
                existing_product.stock = stock_int
                if imagen_url is not None:
                    existing_product.imagen_url = imagen_url
                if descripcion is not None:
                    existing_product.descripcion = descripcion

                session.add(existing_product)
                actualizadas += 1
            else:
                # ── INSERT
                new_product = Product(
                    sku=sku,
                    nombre=nombre,
                    categoria=categoria,
                    precio=precio_dec,
                    stock=stock_int,
                    imagen_url=imagen_url,
                    descripcion=descripcion,
                )
                session.add(new_product)
                insertadas += 1

        # Confirmar todos los cambios válidos en la base de datos
        session.commit()

        return ImportResult(
            leidas=total_leidas,
            insertadas=insertadas,
            actualizadas=actualizadas,
            rechazadas=rechazadas,
            errores=errores,
        )


__all__ = [
    "ImportService",
]
